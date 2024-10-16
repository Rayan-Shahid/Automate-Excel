from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
import os

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get the files from the user
        source_file = request.files['source_file']
        target_file = request.files['target_file']
        source_cell = request.form['source_cell']
        target_column = request.form['target_column']

        if source_file and target_file:
            source_filename = os.path.join(app.config['UPLOAD_FOLDER'], source_file.filename)
            target_filename = os.path.join(app.config['UPLOAD_FOLDER'], target_file.filename)

            # Save the uploaded files
            source_file.save(source_filename)
            target_file.save(target_filename)

            try:
               # procesing files
                copy_data(source_filename, target_filename, source_cell, target_column)
                flash('Data has been successfully copied!', 'success')
            except Exception as e:
               
                flash(f'Error: {str(e)}', 'danger')

            return redirect(url_for('index'))

    return render_template('index.html')

def copy_data(source_path, target_path, source_cell, target_column):
    source_workbook = openpyxl.load_workbook(source_path, data_only=True) 
    target_workbook = openpyxl.load_workbook(target_path)

    target_main_sheet = target_workbook.active

    # Define a dictionary to map sheet names to the corresponding rows in the main sheet
    sheet_name_to_row = {
        "BJ": 3,
        "NP1": 4,
        "NP2": 5,
        "NP3": 6,
        "NP4": 7,
        "NP5": 8,
        "NP7": 9,
        "NP10": 10,
        "NP12": 11,
        "NP13": 12,
        "NP14": 13,
        "NP15": 14,
        "NP16": 15,
        "NP17": 16,
        "NP19": 17,
        "NP20": 18,
        "NP21": 19,
        "NP22": 20,
        "NP24": 21,
        "NP25": 22,
        "NP26": 23,
    }

    for sheet_name, row in sheet_name_to_row.items():
        if sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            try:
                value_to_copy = source_sheet[source_cell].value
                print(f"Copying value from {sheet_name} {source_cell}: {value_to_copy}")

                target_main_sheet[f'{target_column}{row}'] = value_to_copy
                print(f"Pasting value into {target_column}{row} in target sheet.")

            except KeyError as e:
                print(f"Invalid cell reference: {e}")
                raise ValueError(f"Invalid cell reference '{source_cell}' in sheet '{sheet_name}'.")

    target_path2 = os.path.join(app.config['UPLOAD_FOLDER'], 'updated.xlsx')
    target_workbook.save(target_path2)
    print(f"Updated workbook saved as: {target_path2}")


if __name__ == '__main__':
    app.run(debug=True)
